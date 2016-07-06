"""
Created by hatlonely on 2016/07/01.
Copyright © 2016年 hatlonely. All rights reserved.

这是我帮公司同事写的一个程序，她每个月要从考勤机上导出考勤的原始数据（xls格式），
然后手动汇总统计考勤数据，包括每天的上下班时间，迟到次数，加班时长，
总共3000多条考勤记录，这些工作枯燥乏味，而且很费时间，而且还不一定准确，
所以有了这个程序，这个程序从原始的考勤记录数据中去统计汇总，生成一个xlsx的考勤表

输入：考勤的原始数据，input.xls
输出：考勤明细.xlsx
"""
#!/usr/bin/env python3

from openpyxl import Workbook
import xlrd
import datetime
import re

# 工作日
workdayset = [datetime.datetime(2016, 6, 12)]
# 节假日
hoildayset = [datetime.datetime(2016, 6, 9), datetime.datetime(2016, 6, 10)]

# 别名映射表
# 有些人的记录对应了多个名字，通过这里配置，将不同的编号映射到同一个人身上
alias_map = {
    # '27': 'somebody',
    # '307': 'somebody',
}

# 日期范围,从excel文件中获取
# 自动从输入文件中读取起始和截止日期
daterange = []

# 读取excel文件中的数据，生成名字和记录集合的映射
def readinput():
    name_records_map = {}
    datetimeset = set()
    xls = xlrd.open_workbook('input.xls')
    sheet = xls.sheets()[0]
    for i in range(0, sheet.nrows):
        # 第3列是时间，确认这个格式是时间，否则会出异常，一般是表头
        if sheet.cell(i, 3).ctype != 3:
            continue
        name = sheet.cell(i, 1).value
        # 如果名字是某个人的别名，将这个记录合并到这个人身上
        if name in alias_map:
            name = alias_map[name]
        # 如果名字全是数字，已经不知道这个数字对应的是谁的记录，直接忽略
        if re.match(r'[0-9]+', name):
            continue
        ddtt = xlrd.xldate.xldate_as_datetime(sheet.cell(i, 3).value, xls.datemode)
        datetimeset.add(ddtt)
        if name not in name_records_map:
            name_records_map[name] = []
        name_records_map[name].append(ddtt)
    date1 = min(datetimeset)    # 起始日期
    date2 = max(datetimeset)    # 结束日期
    daterange.append(datetime.datetime(date1.year, date1.month, date1.day))
    daterange.append(datetime.datetime(date2.year, date2.month, date2.day))
    return name_records_map

# 分析数据，获取日期到记录的映射
def analysis_date(name_records_map):
    name_date_records_map = {}
    for name in name_records_map:
        name_date_records_map[name] = {}
        records = name_records_map[name]
        for record in records:
            # 向后推移4个小时,处理那种12点之后打卡的异常
            delta = datetime.timedelta(hours=4)
            ddtt = record - delta
            ddtt = datetime.datetime(ddtt.year, ddtt.month, ddtt.day)
            if ddtt not in name_date_records_map[name]:
                name_date_records_map[name][ddtt] = []
            name_date_records_map[name][ddtt].append(record)
    return name_date_records_map

# 分析数据,获取每天的打卡记录
def analysis_work(name_date_records_map):
    name_date_work_map = {}
    for name in name_date_records_map:
        name_date_work_map[name] = {}
        name_date_work_map[name]['迟到'] = 0
        name_date_work_map[name]['早退'] = 0
        name_date_work_map[name]['总加班时长'] = datetime.timedelta(hours=0)
        name_date_work_map[name]['工作日加班时长'] = datetime.timedelta(hours=0)
        name_date_work_map[name]['节假日加班时长'] = datetime.timedelta(hours=0)
        date_records = name_date_records_map[name]
        for ddtt in date_records:
            name_date_work_map[name][ddtt] = {}
            records = date_records[ddtt]
            name_date_work_map[name][ddtt]['上班'] = min(records)
            name_date_work_map[name][ddtt]['下班'] = max(records)
            name_date_work_map[name][ddtt]['加班'] = name_date_work_map[name][ddtt]['下班'] - name_date_work_map[name][ddtt]['上班']
            # 工作日，加班时长 = 下班时间 - 19:30
            if ddtt not in hoildayset and (ddtt.weekday() < 5 or ddtt in workdayset):
                after_work_time = datetime.datetime(ddtt.year, ddtt.month, ddtt.day, 19, 30)
                if name_date_work_map[name][ddtt]['下班'] > after_work_time:
                    name_date_work_map[name][ddtt]['加班'] = name_date_work_map[name][ddtt]['下班'] - after_work_time
                else:
                    name_date_work_map[name][ddtt]['加班'] = datetime.timedelta(hours=0)
                name_date_work_map[name]['工作日加班时长'] += name_date_work_map[name][ddtt]['加班']
            else:
                # 节假日，加班时长 = 上班时间 - 下班时间
                name_date_work_map[name]['节假日加班时长'] += name_date_work_map[name][ddtt]['加班']
            name_date_work_map[name]['总加班时长'] += name_date_work_map[name][ddtt]['加班']
            # 异常处理，如果一天只打了一次卡，时间没有超过12点就算上班打卡，否则算下班打卡，另一个不填，算考勤异常
            if name_date_work_map[name][ddtt]['上班'] == name_date_work_map[name][ddtt]['下班']:
                if name_date_work_map[name][ddtt]['上班'].hour < 12:
                    name_date_work_map[name][ddtt]['下班'] = None
                else:
                    name_date_work_map[name][ddtt]['上班'] = None
            else:
                # 上班时间超过 10:30 算迟到
                if name_date_work_map[name][ddtt]['上班'] > ddtt + datetime.timedelta(hours=10, minutes=30):
                    name_date_work_map[name]['迟到'] += 1
                # 下班时间在 18:30 之前，其实这应该算早退，
                # 但是考虑可能是下午出去打卡，而下班忘记打卡，所以这里算作异常
                if name_date_work_map[name][ddtt]['下班'] < ddtt + datetime.timedelta(hours=18, minutes=30):
                    name_date_work_map[name][ddtt]['下班'] = None
    return name_date_work_map

# 生成考勤表
def kaoqinoutput(name_date_work_map):
    wb = Workbook()
    ws = wb.active
    ws.print_options.horizontalCentered = True

    # 生成日期
    dates = []
    ddtt = daterange[0]
    while ddtt != daterange[1]:
        if ddtt not in hoildayset and (ddtt.weekday() < 5 or ddtt in workdayset):
            dates.append(ddtt)
        ddtt += datetime.timedelta(days=1)
    dates.append(ddtt)
    # 填充日期列
    ws.cell(column=1, row=1, value='时间')
    for i in range(0, len(dates)):
        ddtt = dates[i]
        ws.cell(column=2 * i + 2, row=1, value=ddtt.strftime('%Y-%m-%d'))
        ws.merge_cells(start_row=1, start_column=2 * i + 2, end_row=1, end_column=2 * i + 3)
    ws.cell(column=2 * len(dates) + 2, row=1, value='总加班时长')
    ws.cell(column=2 * len(dates) + 3, row=1, value='工作日加班时长')
    ws.cell(column=2 * len(dates) + 4, row=1, value='节假日加班时长')
    ws.cell(column=2 * len(dates) + 5, row=1, value='迟到')
    # ws.cell(column=2 * len(dates) + 6, row=1, value='早退')
    # 填充状态列
    ws.cell(column=1, row=2, value='状态')
    for i in range(0, len(dates)):
        ws.cell(column=2 * i + 2, row=2, value='上班')
        ws.cell(column=2 * i + 3, row=2, value='下班')
    # 填充数据列
    row = 2
    for name in sorted(name_date_work_map.keys()):
        row += 1
        ws.cell(column=1, row=row, value=name)
        for i in range(0, len(dates)):
            ddtt = dates[i]
            date_work_map = name_date_work_map[name]
            if ddtt in date_work_map:
                if date_work_map[ddtt]['上班']:
                    ws.cell(column=2 * i + 2, row=row, value=date_work_map[ddtt]['上班'].strftime('%H:%M:%S'))
                if date_work_map[ddtt]['下班']:
                    ws.cell(column=2 * i + 3, row=row, value=date_work_map[ddtt]['下班'].strftime('%H:%M:%S'))
        # 加班时长计算方法，加班时长天数 * 24 / 7.5，即每天按7.5天计算
        ws.cell(column=2 * len(dates) + 2, row=row,value=name_date_work_map[name]['总加班时长'].days * 24 / 7.5 + name_date_work_map[name]['总加班时长'].seconds / 3600.0 / 7.5)
        ws.cell(column=2 * len(dates) + 3, row=row, value=name_date_work_map[name]['工作日加班时长'].days * 24 / 7.5 + name_date_work_map[name]['工作日加班时长'].seconds / 3600.0 / 7.5)
        ws.cell(column=2 * len(dates) + 4, row=row, value=name_date_work_map[name]['节假日加班时长'].days * 24 / 7.5 + name_date_work_map[name]['节假日加班时长'].seconds / 3600.0 / 7.5)
        ws.cell(column=2 * len(dates) + 5, row=row, value=name_date_work_map[name]['迟到'])
        # ws.cell(column=2 * len(dates) + 6, row=row, value=name_date_work_map[name]['早退'])
    wb.save('考勤明细.xlsx')

def main():
    data = analysis_work(analysis_date(readinput()))
    kaoqinoutput(data)

if __name__ == '__main__':
    main()

